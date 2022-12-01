from openpyxl import load_workbook

workbook = load_workbook(filename="mapping-example-table.xlsx")
print(workbook.sheetnames)

sheet = workbook.active
print(sheet)

print(sheet.title)